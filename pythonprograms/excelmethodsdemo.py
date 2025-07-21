from openpyxl import load_workbook, Workbook

def create_excel_file(file,data):
    wb = Workbook()
    sheet = wb.active
    sheet.title="Bank Data"

    for row in data:
        sheet.append(row)

    wb.save(file)
    print("Data written to exel file ...!")


def read_from_excel(file):
    wb = load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        print(row)


def update_cell_data(filename, row, column, value):
    wb = load_workbook(filename)
    sheet = wb.active
    #oldvalue = sheet.cell(row=row, column=column).value
    sheet.cell(row=row,column=column).value=value
    wb.save(filename)
    print("Updated data")


def add_new_row(filename, rdata):
    wb= load_workbook(filename)
    sheet = wb.active
    sheet.append(rdata)
    wb.save(filename)
    print("Data added to sheet")



def delete_row(filename,num):
    wb = load_workbook(filename)
    sheet = wb.active
    sheet.delete_rows(num)
    wb.save(filename)
    print("Row data deleted..!")


def main():
    bankdata = [
        ["tid", "mode", "type", "amount"],
        [9001, "cash", "deposit", 10000],
        [9002, "cheque", "withdrawal", 10000],
    [9003, "cash", "deposit", 20000],
    [9004, "upi", "transfer", 4000]
    ]
    filename = input("Enter the filename: ")
    #  create_excel_file(filename,bankdata)
    #   read_from_excel(filename)
    #num = int(input("Enter the row number to delete: "))
    #  delete_row(filename,num)
    #row = int(input("Enter the row number: "))
    #column = int(input("Enter the column number: "))
    #value = int(input("Enter the value: "))
    #update_cell_data(filename, row, column, value)
    #read_from_excel(filename)
    rdata = [9006,'cash','deposit',20000]
    add_new_row(filename,rdata)
    read_from_excel(filename)


if __name__ == '__main__':
    main()
