from openpyxl import load_workbook
from employeedao import EmployeeDao
from employee import Employee


class EmployeeService:

   def __init__(self):
      self.dao  = EmployeeDao()


   def clean_employee_data(self, employees):
      pass

   def load_employee_data_todb(self,employees):
      self.dao.add_employee(employees)

   def read_from_excel(filename):
      empservice = EmployeeService()
      try:
         wb = load_workbook(filename)
         sheet = wb.active
         employees = []
         for row in sheet.iter_rows(min_rows=2, values_only=True):
            empid, name, age, dept, notes, salary, remarks = row
            employee = Employee(empid, name, age, dept, notes, salary, remarks)
            employees.append(employee)

         return empservice.clean_employee_data(employees)
      except Exception as e:
         print(e)


